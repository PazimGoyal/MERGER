import pandas as pd
import numpy as np
import shutil
import os
import json
from datetime import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

with open('BACKUP/fields.json') as f:
    data = json.load(f)
columns = data['files']['COLUMNS']


def delete_accounting():
    os.remove(data['files']['accounting_file'])
    # pass


def check_vch(vchno):
    try:
        df = pd.read_csv('BACKUP/vchrs.csv', header=None)
        if vchno in list(df[0]):
            return False
        else:
            return True
    except:
        return True


def add_ch(vchno):
    with open('BACKUP/vchrs.csv', 'a') as f:
        pd.DataFrame([vchno]).to_csv(f, header=False, index=False)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, set_width=False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        if 'header' in to_excel_kwargs:
            to_excel_kwargs.pop('header')

        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    if set_width:
        size_dict = {1: 15, 3: 15, 4: 35, 5: 15, 6: 15, 7: 20, 8: 10, 9: 10, 10: 10, 12: 30, 14: 20, 15: 20, 16: 20}
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            to_excel_kwargs.pop('header')

        else:
            dim_holder = DimensionHolder(worksheet=ws)
            for i, j in size_dict.items():
                dim_holder[get_column_letter(i)] = ColumnDimension(ws, min=i, max=i, width=j)
            ws.column_dimensions = dim_holder
            ws.protection.sheet = True

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def make_backup(path='BACKUP/BACKUP'):
    shutil.copy2('BANKS.xlsx', path + '/BANKS.xlsx')
    shutil.copy2('FirmsSorted.xlsx', path + '/FirmsSorted.xlsx')
    shutil.copy2('FRT.CAL.xlsx', path + '/FRT.CAL.xlsx')


def file_merge():
    f = open(date_stored, 'w+')
    f.write(last_modified)
    f.close()
    make_backup()
    try:
        df_accounting = pd.read_excel(data['files']['accounting_file'], header=None, dtype=str)
        df_trucking = pd.read_excel(data['files']['trucking_file'], sheet_name=data['files']['sheet_name'], dtype=str)
    except Exception as e:
        print(e)
        raise e
    list_all_rows = list()
    data_acc = data['account']
    data_truck = data['truck']
    start = data_acc['START']
    empty_count = 0
    max_size = df_accounting.shape[0]

    firms = {'GKF': [], 'HTC': [], 'BRT': [], 'UAE': [], 'PUN': [], 'KAI': [], 'OTHERS': []}
    banks = {"SBI": [], "OTHERS": []}

    empty_truck_df = pd.DataFrame({'SR.NO': '', 'TRUCK NO': '', 'BANK ACCOUNT NO': '', 'IFSC CODE': '', 'BANK NAME': '',
                                   'OWNER NAME': '', 'MOBILE NO': ''}, index=[0]
                                  )

    vhno = df_accounting.iat[data_acc['VCH.NO'][1], data_acc['VCH.NO'][0]]
    check_vch_exist = check_vch(vhno)
    date_vch = df_accounting.iat[data_acc['VCH.DATE'][1], data_acc['VCH.DATE'][0]]
    rh = df_accounting.iat[data_acc['REF'][1], data_acc['REF'][0]]
    credit = 'ERROR'
    if check_vch_exist:
        while start <= max_size:
            try:
                list_row = []
                data_row = df_accounting.loc[start]
                Truck_no = data_row.iat[data_acc['TRUCK']]
                if data_row.iat[data_acc['END']] == 'On Account of :' or data_row.iat[data_acc['END']] == 'To':
                    credit = data_row.iat[1]
                    break
                if data_row.empty:
                    empty_count += 1
                    if empty_count > 3:
                        break

                if Truck_no is np.nan:
                    Truck_no = ''
                data_truck_row = df_trucking.loc[df_trucking[data_truck['TRUCK']] == Truck_no]
                if data_truck_row.empty:
                    data_truck_row = empty_truck_df
                list_row.append(vhno)
                list_row.append(rh)
                list_row.append(date_vch)
                chalan = data_row.iat[data_acc['CHALLAN']]
                list_row.append(chalan)
                list_row.append(data_row.iat[data_acc['DATE']])
                list_row.append(Truck_no)
                list_row.append(data_truck_row[data_truck['NAME']].values[0])
                list_row.append(data_truck_row[data_truck['AC.NO']].values[0])
                list_row.append(data_truck_row[data_truck['IFSC']].values[0])
                bank = data_truck_row[data_truck['BANK']].values[0]
                list_row.append(bank)
                list_row.append(data_row.iat[data_acc['AMOUNT']])
                list_row.append(data_row.iat[data_acc['DEST']])
                list_row.append(data_row.iat[data_acc['QTY']])
                list_row.append(data_truck_row[data_truck['MOB.NO']].values[0])
                list_row.append(data_row.iat[data_acc['GNO']])
                list_all_rows.append(list_row)
                types = chalan.split('/')

                firm = firms.get(types[2])
                if types[1] != 'RH' or firm is None:
                    firm = firms.get("OTHERS")
                firm.append(list_row)

                if bank == 'STATE BANK OF INDIA' or bank == 'SBI':
                    bank_l = banks.get('SBI')
                else:
                    bank_l = banks.get("OTHERS")

                bank_l.append(list_row)


            except Exception as e:
                print(e)
            start += 1
    else:
        print("This Voucher Number is already imported, Kindly change voucher number, Press any key to exit")
        input()

    final = [""] * 16
    final[5] = "Total Account VCH:"
    final[6] = df_accounting.iat[data_acc['TOTAL'][1], data_acc['TOTAL'][0]]
    final[9] = "Total Calculated:"
    final[10] = ''

    now = dt.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    empt = [""] * 16
    empt[0] = '-'
    list_all_rows.append(empt)
    final[2] = "Last Modified : " + dt_string
    list_all_rows.append(final)
    list_all_rows.append(empt)
    list_all_rows.append(empt)
    new_df = pd.DataFrame(list_all_rows, columns=columns).astype(str)
    new_df['Credit'] = credit
    new_df['Credit']
    new_df['Credit'][-4:] = ""
    total = pd.to_numeric(new_df['AMOUNT']).sum()
    new_df.loc[(new_df.shape[0] - 3), 'AMOUNT'] = total
    append_df_to_excel(data['files']['calculation_file'], new_df, set_width=True, header=None, index=False)
    final[5] = ""
    final[6] = ""
    add_ch(vhno)

    for i, j in firms.items():
        if j:
            j.append(empt)
            j.append(final)
            j.append(empt)
            new_df = pd.DataFrame(j, columns=columns).astype(str)
            new_df['Credit'] = credit
            new_df['Credit'][-3:] = ""
            total = pd.to_numeric(new_df['AMOUNT']).sum()
            new_df.loc[(new_df.shape[0] - 2), 'AMOUNT'] = total
            append_df_to_excel('FirmsSorted.xlsx', new_df, set_width=True, header=None, index=False, sheet_name=i)

    for i, j in banks.items():
        if j:
            j.append(empt)
            j.append(final)
            j.append(empt)
            new_df = pd.DataFrame(j, columns=columns).astype(str)
            new_df['Credit'] = credit
            new_df['Credit'][-3:] = ""
            total = pd.to_numeric(new_df['AMOUNT']).sum()
            new_df.loc[(new_df.shape[0] - 2), 'AMOUNT'] = total
            append_df_to_excel('BANKS.xlsx', new_df, set_width=True, header=None, index=False, sheet_name=i)


    # delete_accounting()


date_stored = 'BACKUP/last.txt'
f = open(date_stored, 'r')
stored = f.readline()
last_modified = str(os.path.getmtime(data['files']['accounting_file']))
if last_modified == stored:
    print("ACCOUNTING File was not modified since last run. Do you Still want to continue....")
    reply = input("Press Y if you want to continue or any other key to exit   ")
    if reply == 'Y' or reply == 'y':
        file_merge()
else:
    file_merge()
